import copy
import os
import re
import csv
from openpyxl import load_workbook

def unmerge_cells(workbook):
    # Iterate over sheets in workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        merged_cell_ranges = copy.deepcopy(workbook[sheet_name].merged_cells.ranges)

        # Iterate over merged cells
        for merged_cell_range in merged_cell_ranges:
            min_row = merged_cell_range.min_row
            max_row = merged_cell_range.max_row
            min_col = merged_cell_range.min_col
            max_col = merged_cell_range.max_col

            for row in sheet.iter_rows(min_row,max_row,min_col,max_col):
                for cell in row:
                    found_value = None
                    if cell.value:
                        value = cell.value
                        found_value = True
                        break
                if found_value:
                    sheet.unmerge_cells(str(merged_cell_range))
                    for cell in row:
                       sheet[cell.coordinate].value = value

def format_header_row(workbook, header_row):
    # Iterate over sheets in workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(header_row - 1, header_row - 1, 1, sheet.max_column):
            for cell in row:
                cell_text = cell.value
                if cell_text is not None:
                    clean_text = cell_text.split(' ')[0]
                    sheet[cell.coordinate].value = clean_text
        for row in sheet.iter_rows(header_row, header_row, 1, sheet.max_column):
            for cell in row:
                cell_text = cell.value
                if cell_text is not None:
                    clean_text = cell_text.split('\n')[0]
                    clean_text = re.split(r"\(.*?\)", clean_text)[0]
                    clean_text = clean_text.replace("#", "")
                    clean_text = clean_text.replace('%', 'pct')
                    clean_text = clean_text.rstrip()
                    clean_text = clean_text.replace(" ", "_")
                    clean_text = clean_text.replace(".", "")
                    clean_text = clean_text.replace("-", "_")
                    sheet[cell.coordinate].value = clean_text

def clean_data(workbook):
    clean_data = [] # this will be a list of dictionaries
    new_row = {}  # dictionary containing information about one student, one assessment
        #TODO: find non-unique rows of ids and message error
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        max_row = sheet.max_row

        # find all values in row student ID, create new dictionary for each column with a value
        for row in sheet.iter_rows(4, max_row):
            for cell in row:
                value = cell.value

                # value > 0 skips composite scores for assessments that haven't been entered
                # cell.column > 3 skips the cells first 3 columns (Teacher, Student, ID)
                # from being counted as a new data entry
                # not None skips empty cells
                if value is not None and cell.column > 3 and value > 0:
                    new_row['MO_ID'] = int(sheet.cell(cell.row, column=3).value)
                    new_row['teacher'] = sheet.cell(row = cell.row, column = 1).value
                    new_row['grade'] = sheet.title
                    new_row['assessment period'] = sheet.cell(row = 2, column = cell.column).value
                    new_row['assessment'] = sheet.cell(row = 3, column=cell.column).value
                    new_row['score'] = value
                    clean_data.append(new_row)
                    new_row = {}
    return clean_data

def write_to_csv(data, output_csv_path):
    # Get the keys from the first dictionary to use as column headers
    fieldnames = data[0].keys()

    # Write the list of dictionaries to the CSV file
    with open(output_csv_path, 'w', newline='') as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()  # Write the column headers
        writer.writerows(data)  # Write the data rows






# Specify the header row index

# TODO add error messages to help see where the format of the xlsx may be causing trouble.

def main():

    header_row = 3  # Specify the header row index

    # Provide the file paths for the Excel and CSV files
    input_file_path = input('Enter the path of the .xlsx file: ')
    print(input_file_path)

    if os.path.exists(input_file_path):
        print('The input file exists')
    else:
        print('The specified input file path does *not* exists')

    output_file_path = input("Enter the full path and filename to save the output: ")
    print(output_file_path)


    workbook = load_workbook(input_file_path, data_only=True)
    unmerge_cells(workbook)
    format_header_row(workbook,header_row)
    data = clean_data(workbook)
    write_to_csv(data, output_file_path)
    print('CSV has been saved as: ', output_file_path)

if __name__ == "__main__":
    main()

# this_workbook.save(output_xlsx_path)